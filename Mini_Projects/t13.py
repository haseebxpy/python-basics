from openpyxl import Workbook
workbook=Workbook()
sheet=workbook.active
sheet.append(["Name","Department","Salary"])
sheet.append(["Haseeb","IT","500"])
sheet.append(["Ali","HR","400"])
sheet.append(["Sara","IT","600"])
workbook.save("input13.xlsx")
from openpyxl import load_workbook
workbook= load_workbook("input13.xlsx")
sheet=workbook.active
totalsalary=0
itdepsalary=0
for row in sheet.iter_rows(min_row=2,values_only=True):
    name,department,salary=row
    totalsalary+=int(salary)
    if department.strip().lower()=="it":
        itdepsalary+=int(salary)
print(f"Total Salary : {totalsalary}")
print(f"IT Department Salary : {itdepsalary}")


    